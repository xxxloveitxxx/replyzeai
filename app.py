import os
import time
import base64
import requests
import smtplib
import imaplib
import ssl
import io
from flask import abort, Flask, render_template, request, redirect, jsonify, make_response, url_for
from datetime import date, datetime, timezone, timedelta
from email.mime.text import MIMEText
from supabase import create_client, Client
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import Flow
from google.oauth2 import id_token
import google.auth.transport.requests as grequests
from fimap import send_email_smtp, fetch_emails_imap
from flask_cors import CORS  
from cryptography.fernet import Fernet
from transaction_autopilot import bp as autopilot_bp
from public import public_bp
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import re
import dns.resolver
import csv
from io import TextIOWrapper
from openpyxl import load_workbook
from collections import defaultdict
from functools import wraps

# ── single Flask app & blueprint registration ──
app = Flask(__name__, template_folder="templates")
CORS(app, resources={r"/connect-smtp": {"origins": "https://replyzeai.vercel.app"}})
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret")

# Rate limiting storage
# Rate limiting storage - fix structure and initialization
demo_rate_limits = defaultdict(lambda: {
    'emails': {'remaining': 20, 'last_reset': datetime.now()},
    'kits': {'remaining': 20, 'last_reset': datetime.now()},
    'leads': {'remaining': 25, 'last_reset': datetime.now()}
})

# Fixed rate limit decorator
def check_rate_limit(resource):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            ip = request.remote_addr
            now = datetime.now()
            
            # Reset limits based on their specific time periods
            if resource == 'emails':
                # Daily reset for emails
                if (now - demo_rate_limits[ip][resource]['last_reset']).days >= 1:
                    demo_rate_limits[ip][resource]['remaining'] = 20
                    demo_rate_limits[ip][resource]['last_reset'] = now
            else:
                # Monthly reset for kits and leads
                if (now - demo_rate_limits[ip][resource]['last_reset']).days >= 30:
                    if resource == 'kits':
                        demo_rate_limits[ip][resource]['remaining'] = 20
                    else:  # leads
                        demo_rate_limits[ip][resource]['remaining'] = 25
                    demo_rate_limits[ip][resource]['last_reset'] = now
            
            # Check if limit is exceeded
            if demo_rate_limits[ip][resource]['remaining'] <= 0:
                return jsonify({"error": f"{resource.capitalize()} limit exceeded"}), 429
            
            # Decrement the counter and proceed
            demo_rate_limits[ip][resource]['remaining'] -= 1
            return f(*args, **kwargs)
        return decorated_function
    return decorator

#--------------------------------------------------------------
@app.route("/signin")
def signin():
    user_id = request.args.get("user_id", "")
    return render_template("signin.html", user_id=user_id)
#--------------------------------------------------------------
app.register_blueprint(autopilot_bp, url_prefix="/autopilot")
app.register_blueprint(public_bp)

# --- Supabase setup ---
SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_ANON_KEY = os.environ["SUPABASE_ANON_KEY"]
SUPABASE_SERVICE_ROLE_KEY = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
supabase: Client = create_client(SUPABASE_URL, SUPABASE_ANON_KEY)
SUPABASE_SERVICE: Client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
# Edge Function base URL *without* trailing slash or endpoint
EDGE_BASE_URL = os.environ.get("EDGE_BASE_URL", "").rstrip("/")
ENCRYPTION_KEY = os.environ["ENCRYPTION_KEY"].encode()  # 32-url-safe-base64 bytes
fernet = Fernet(ENCRYPTION_KEY)
# Retry configuration for calling the Edge Function
MAX_RETRIES = 5
RETRY_BACKOFF_BASE = 2

# Define follow-up sequence (days after initial contact)
FOLLOW_UP_SEQUENCE = [
    {"delay_days": 0, "name": "Immediate Follow-up"},
    {"delay_days": 1, "name": "Day 1 Follow-up"},
    {"delay_days": 3, "name": "Day 3 Follow-up"},
    {"delay_days": 7, "name": "Day 7 Follow-up"},
    {"delay_days": 14, "name": "Day 14 Follow-up"},
    {"delay_days": 30, "name": "Day 30 Follow-up"},
]

#----------------------------------------------------------------------------
def get_smtp_creds(user_id: str):
    """Return decrypted (email, app_password) or (None, None)."""
    try:
        resp = supabase.from_("profiles").select("smtp_email, smtp_enc_password").eq("id", user_id).single().execute()
        
        # Check if response has data
        if not resp.data:
            app.logger.warning(f"No SMTP credentials found for user {user_id}")
            return None, None
            
        # Check if password exists
        if not resp.data.get("smtp_enc_password"):
            app.logger.warning(f"No SMTP password found for user {user_id}")
            return None, None
            
        enc_pwd = resp.data["smtp_enc_password"].encode()
        try:
            pwd = fernet.decrypt(enc_pwd).decode()
        except Exception as e:
            app.logger.error(f"Failed to decrypt password for user {user_id}: {str(e)}")
            return None, None
            
        return resp.data["smtp_email"], pwd
        
    except Exception as e:
        app.logger.error(f"Error retrieving SMTP credentials for user {user_id}: {str(e)}")
        return None, None

# ---------------------------------------------------------------------------
def call_edge(endpoint_path: str, payload: dict, return_response: bool = False):
    url = f"{EDGE_BASE_URL}{endpoint_path}"
    app.logger.info(f"🔗 call_edge → URL: {url}")
    app.logger.info(f"🔗 call_edge → Payload: {payload}")

    headers = {
        "Authorization": f"Bearer {SUPABASE_SERVICE_ROLE_KEY}",
        "apikey":        SUPABASE_SERVICE_ROLE_KEY,
        "Content-Type":  "application/json"
    }

    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.post(url, json=payload, headers=headers, timeout=120)
            app.logger.info(f"↩️  Response [{resp.status_code}]: {resp.text}")

            if resp.status_code == 200:
                if return_response:
                    return resp
                else:
                    return True
            elif resp.status_code == 429:
                wait = RETRY_BACKOFF_BASE ** attempt
                app.logger.warning(f"[{endpoint_path}] Rate‐limited, retry {attempt+1}/{MAX_RETRIES} after {wait}s")
                time.sleep(wait)
                continue
            else:
                app.logger.error(f"[{endpoint_path}] Failed ({resp.status_code}): {resp.text}")
                if return_response:
                    return resp
                else:
                    return False
        except requests.RequestException as e:
            wait = RETRY_BACKOFF_BASE ** attempt
            app.logger.error(f"[{endpoint_path}] Exception: {e}, retrying in {wait}s")
            time.sleep(wait)
    app.logger.error(f"[{endpoint_path}] Exceeded max retries.")
    if return_response:
        return None
    else:
        return False

# ── Routes ──
#-----------------------------------------------


# Add this near the top of your app.py after creating the Flask app
@app.template_filter('format_date')
def format_date_filter(value):
    if not value:
        return ""
    try:
        # Try to parse the date string
        date_obj = datetime.fromisoformat(value.replace('Z', '+00:00'))
        return date_obj.strftime("%b %d, %Y %I:%M %p")
    except:
        return value



# Add this function near your other helper functions
def verify_smtp_connection(user_id: str) -> dict:
    """
    Test SMTP/IMAP connection and return status
    Returns: {"status": "valid"|"invalid", "message": str}
    """
    try:
        # Get SMTP credentials
        smtp_email, app_password = get_smtp_creds(user_id)
        if not smtp_email or not app_password:
            return {"status": "invalid", "message": "No SMTP credentials found"}
        
        # Get server details
        resp = supabase.from_("profiles").select(
            "smtp_host, imap_host"
        ).eq("id", user_id).single().execute()
        
        if not resp.data:  # Changed from: if resp.error or not resp.data:
            return {"status": "invalid", "message": "Could not retrieve server details"}
        
        server_details = resp.data
        smtp_host = server_details.get("smtp_host", "smtp.gmail.com")
        smtp_port = 587
        imap_host = server_details.get("imap_host", "imap.gmail.com")
        imap_port = 993
        
        # Test SMTP connection
        smtp_working = False
        imap_working = False
        smtp_error = None
        imap_error = None
        
        try:
            context = ssl.create_default_context()
            with smtplib.SMTP(smtp_host, smtp_port) as server:
                server.ehlo()
                server.starttls(context=context)
                server.ehlo()
                server.login(smtp_email, app_password)
                smtp_working = True
        except Exception as e:
            smtp_error = str(e)
            app.logger.error(f"SMTP test failed for {user_id}: {smtp_error}")
        
        # Test IMAP connection
        try:
            with imaplib.IMAP4_SSL(imap_host, imap_port) as server:
                server.login(smtp_email, app_password)
                imap_working = True
        except Exception as e:
            imap_error = str(e)
            app.logger.error(f"IMAP test failed for {user_id}: {imap_error}")
        
        # Update database with connection status
        status = "valid" if smtp_working and imap_working else "invalid"
        supabase.table("profiles").update({
            "email_connection_status": status,
            "connection_checked_at": datetime.now(timezone.utc).isoformat()
        }).eq("id", user_id).execute()
        
        if smtp_working and imap_working:
            return {"status": "valid", "message": "SMTP and IMAP connections successful"}
        elif smtp_working:
            return {"status": "partial", "message": f"SMTP working but IMAP failed: {imap_error}"}
        elif imap_working:
            return {"status": "partial", "message": f"IMAP working but SMTP failed: {smtp_error}"}
        else:
            return {"status": "invalid", "message": f"Both SMTP and IMAP failed. SMTP: {smtp_error}, IMAP: {imap_error}"}
    
    except Exception as e:
        app.logger.error(f"Verification error for {user_id}: {str(e)}")
        return {"status": "invalid", "message": f"Verification error: {str(e)}"}


@app.route('/detect_email_settings', methods=['GET', 'POST'])
def detect_email_settings():
    if request.method == 'GET':
        # Handle GET request (for testing or direct browser access)
        email = request.args.get('email')
    else:
        # Handle POST request
        email = request.form.get('email')
    
    if not email:
        return jsonify({"error": "Email is required"}), 400
    
    try:
        settings = detect_email_provider(email)
        return jsonify(settings)
    except Exception as e:
        app.logger.error(f"Error detecting email settings: {e}")
        # Fall back to Gmail settings
        return jsonify({
            "smtp_host": "smtp.gmail.com",
            "smtp_port": 465,
            "imap_host": "imap.gmail.com",
            "imap_port": 993
        })



def detect_email_provider(email):
    """
    Detect email provider based on domain without network calls
    """
    domain = email.split('@')[-1].lower()
    
    provider_map = {
        "gmail.com": {
            "smtp_host": "smtp.gmail.com",
            "smtp_port": 465,
            "imap_host": "imap.gmail.com",
            "imap_port": 993
        },
        "outlook.com": {
            "smtp_host": "smtp-mail.outlook.com",
            "smtp_port": 587,
            "imap_host": "outlook.office365.com",
            "imap_port": 993
        },
        "yahoo.com": {
            "smtp_host": "smtp.mail.yahoo.com",
            "smtp_port": 465,
            "imap_host": "imap.mail.yahoo.com",
            "imap_port": 993
        },
        "aol.com": {
            "smtp_host": "smtp.aol.com",
            "smtp_port": 465,
            "imap_host": "imap.aol.com",
            "imap_port": 993
        },
        # Add more providers as needed
    }
    
    # Return settings for known providers, or default to Gmail
    return provider_map.get(domain, provider_map["gmail.com"])

# Add this route for checking connection status
@app.route("/check_email_connection")
def check_email_connection():
    user_id = request.args.get("user_id")
    if not user_id:
        return jsonify({"status": "error", "message": "Missing user_id"}), 400
    
    result = verify_smtp_connection(user_id)
    return jsonify(result)

# Add this function to check connection before allowing email features
def require_valid_email_connection(user_id):
    """Check if user has valid email connection, abort if not"""
    # First check if they have Gmail OAuth
    try:
        toks = supabase.table("gmail_tokens").select("credentials").eq("user_id", user_id).execute().data
        if toks and toks[0]:
            cd = toks[0]["credentials"]
            creds = Credentials(
                token=cd["token"],
                refresh_token=cd["refresh_token"],
                token_uri=cd["token_uri"],
                client_id=cd["client_id"],
                client_secret=cd["client_secret"],
                scopes=cd["scopes"],
            )
            if not creds.expired or (creds.expired and creds.refresh_token):
                return True  # Gmail OAuth is valid
    except Exception:
        pass
    
    # Check SMTP/IMAP connection
    smtp_status = verify_smtp_connection(user_id)
    if smtp_status["status"] != "valid":
        abort(403, "Email connection not verified. Please check your settings.")
    
    return True


#-------------------------------------------------
from flask import url_for

@app.route("/")
def home():
    """
    Just redirect to /dashboard, passing along user_id if any.
    """
    user_id = request.args.get("user_id", "")
    # Redirect to /dashboard?user_id=<...> (blank if none)
    return redirect(f"/dashboard?user_id={user_id}")



@app.route("/dashboard")
def dashboard():
    user_id = request.args.get("user_id", "").strip()

    # ── GUEST DEFAULTS ──
    name            = "Guest"
    ai_enabled      = False
    generate_leases = False
    emails_sent     = 0
    time_saved      = 0
    show_reconnect  = False
    revenue         = 0
    revenue_change  = 0

    # Ensure these always exist for the template
    kits_generated = 0
    estimated_saved = 0

    if user_id:
        # 1) Load profile
        try:
            resp = (
                supabase.table("profiles")
                         .select("full_name, ai_enabled, generate_leases")
                         .eq("id", user_id)
                         .single()
                         .execute()
            )
            if resp.data:
                name            = resp.data["full_name"]
                ai_enabled      = resp.data["ai_enabled"]
                generate_leases = resp.data["generate_leases"]
        except Exception:
            app.logger.warning(f"dashboard: failed to load profile for {user_id}")

        # 2) Count today's emails
        try:
            today = date.today().isoformat()
            rows  = (
                supabase.table("emails")
                        .select("sent_at")
                        .eq("user_id", user_id)
                        .eq("status", "sent")
                        .execute()
                        .data
                or []
            )
            emails_sent = sum(1 for e in rows if e.get("sent_at","").startswith(today))
            time_saved  = emails_sent * 5.5
        except Exception:
            app.logger.warning(f"dashboard: failed to count emails for {user_id}")

        # 3) Gmail reconnect flag
        try:
            toks = (
                supabase.table("gmail_tokens")
                         .select("credentials")
                         .eq("user_id", user_id)
                         .execute()
                         .data
                or []
            )
            if toks:
                cd = toks[0]["credentials"]
                creds = Credentials(
                    token=cd["token"],
                    refresh_token=cd["refresh_token"],
                    token_uri=cd["token_uri"],
                    client_id=cd["client_id"],
                    client_secret=cd["client_secret"],
                    scopes=cd["scopes"],
                )
                show_reconnect = creds.expired
        except Exception:
            app.logger.warning(f"dashboard: failed to check Gmail token for {user_id}")

        # 4) Count "kits generated" for this user
        kit_rows = (
            supabase.table("transactions")
                     .select("id")
                     .eq("user_id", user_id)
                     .eq("kit_generated", True)
                     .execute()
                     .data
            or []
        )
        kits_generated = len(kit_rows)

        # 5) Compute extra estimated time saved (e.g. 15 min per kit)
        PER_KIT_SAVE_MINUTES = 15
        estimated_saved = kits_generated * PER_KIT_SAVE_MINUTES

    # ── Render dashboard ──
    return render_template(
        "dashboard.html",
        user_id=user_id,
        name=name,
        ai_enabled=ai_enabled,
        generate_leases=generate_leases,
        emails_sent=emails_sent,
        time_saved=time_saved,
        estimated_saved=estimated_saved,
        kits_generated=kits_generated,
        show_reconnect=show_reconnect,
        revenue=revenue,
        revenue_change=revenue_change
    )
#--------------------------------------------------------------------------------------------------------------
@app.route("/dashboard/leads")
def dashboard_leads():
    user_id = _require_user()
    return render_template("partials/leads_funnel.html", user_id=user_id)

# Fix for the search error - update the leads_list function
@app.route("/dashboard/leads/list")
def leads_list():
    user_id = _require_user()
    filter_type = request.args.get("filter", "all")
    search_query = request.args.get("q", "")
    
    # Build query based on filters
    query = supabase.table("leads").select("*").eq("user_id", user_id)
    
    if filter_type != "all":
        query = query.eq("status", filter_type)
    
    # Execute query first to get all results
    try:
        result = query.execute()
        leads = result.data or []
    except Exception as e:
        app.logger.error(f"Error fetching leads: {str(e)}")
        leads = []
    
    # Apply search filter in Python
    if search_query:
        search_lower = search_query.lower()
        leads = [lead for lead in leads if 
                (lead.get("first_name", "").lower().find(search_lower) != -1 or
                 lead.get("last_name", "").lower().find(search_lower) != -1 or
                 lead.get("email", "").lower().find(search_lower) != -1 or
                 lead.get("brokerage", "").lower().find(search_lower) != -1)]
    
    # Calculate funnel counts
    counts = {
        "new": 0,
        "contacted": 0,
        "proposal": 0,
        "closed": 0
    }
    
    try:
        # Get counts for each status
        for status in counts.keys():
            count_result = supabase.table("leads").select("id", count="exact").eq("user_id", user_id).eq("status", status).execute()
            counts[status] = count_result.count or 0
    except Exception as e:
        app.logger.error(f"Error counting leads by status: {str(e)}")
    
    return render_template("partials/leads_list.html", leads=leads, counts=counts, user_id=user_id)


@app.route("/dashboard/leads/search")
def search_leads():
    # Reuse the leads_list function but with search parameters
    return leads_list()

@app.route("/dashboard/leads/<lead_id>/view")
def view_lead(lead_id):
    user_id = _require_user()
    
    try:
        # Get lead details
        lead = supabase.table("leads").select("*").eq("id", lead_id).eq("user_id", user_id).single().execute().data
        
        # Get follow-up history
        follow_ups = supabase.table("lead_follow_ups").select("*").eq("lead_id", lead_id).order("scheduled_at").execute().data or []
        
        return render_template("partials/lead_detail.html", lead=lead, follow_ups=follow_ups, user_id=user_id)
    except Exception as e:
        app.logger.error(f"Error fetching lead details: {str(e)}")
        return "<div class='error'>Error loading lead details: Missing required database columns</div>", 500

@app.route("/dashboard/leads/<lead_id>/update-status", methods=["POST"])
def update_lead_status(lead_id):
    user_id = _require_user()
    new_status = request.form.get("status")
    
    if not new_status:
        return jsonify({"error": "Status is required"}), 400
    
    try:
        # Update lead status
        supabase.table("leads").update({
            "status": new_status,
            "last_updated": datetime.now(timezone.utc).isoformat()
        }).eq("id", lead_id).eq("user_id", user_id).execute()
        
        return "", 204
    except Exception as e:
        app.logger.error(f"Error updating lead status: {str(e)}")
        return jsonify({"error": "Failed to update status"}), 500

# Fix for the lead notes error - update the add_lead_note function
@app.route("/dashboard/leads/<lead_id>/add-note", methods=["POST"])
def add_lead_note(lead_id):
    user_id = _require_user()
    note_content = request.form.get("note")
    
    if not note_content:
        return jsonify({"error": "Note content is required"}), 400
    
    try:
        # First verify the lead exists and belongs to this user
        lead_check = supabase.table("leads").select("id").eq("id", lead_id).eq("user_id", user_id).execute()
        if not lead_check.data:
            return jsonify({"error": "Lead not found or access denied"}), 404
        
        # Add note to lead
        result = supabase.table("lead_notes").insert({
            "lead_id": lead_id,
            "user_id": user_id,
            "content": note_content,
            "created_at": datetime.now(timezone.utc).isoformat()
        }).execute()
        
        # Check if insertion was successful
        if not result.data:
            app.logger.error(f"Note insertion failed: {result}")
            return jsonify({"error": "Failed to add note - no data returned"}), 500
            
        return "", 204
    except Exception as e:
        app.logger.error(f"Error adding lead note: {str(e)}", exc_info=True)
        
        # Check if it's a specific API error
        error_msg = str(e)
        if "foreign key constraint" in error_msg.lower():
            return jsonify({"error": "Invalid lead ID"}), 400
        elif "null value" in error_msg.lower():
            return jsonify({"error": "Missing required fields"}), 400
            
        return jsonify({"error": "Failed to add note"}), 500


@app.route("/dashboard/leads/export")
def export_leads():
    user_id = _require_user()
    filter_type = request.args.get("filter", "all")
    
    try:
        # Build query
        query = supabase.table("leads").select("*").eq("user_id", user_id)
        
        if filter_type != "all":
            query = query.eq("status", filter_type)
        
        leads = query.execute().data or []
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow(["First Name", "Last Name", "Email", "Brokerage", "Service", "City", "Status", "Last Contact"])
        
        # Write data
        for lead in leads:
            writer.writerow([
                lead.get("first_name", ""),
                lead.get("last_name", ""),
                lead.get("email", ""),
                lead.get("brokerage", ""),
                lead.get("service", ""),
                lead.get("city", ""),
                lead.get("status", "new"),
                lead.get("last_contacted_at", "")
            ])
        
        # Prepare response
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename=leads_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        response.headers["Content-type"] = "text/csv"
        
        return response
    except Exception as e:
        app.logger.error(f"Error exporting leads: {str(e)}")
        return jsonify({"error": "Failed to export leads"}), 500
#------------------------------------------------------------------------------------------------------------
@app.route("/dashboard/new_transaction")
def dashboard_new_transaction():
    user_id = request.args.get("user_id") or abort(401)
    return render_template("partials/new_transaction.html", user_id=user_id)
  
@app.route("/dashboard/responded_emails")
def dashboard_responded_emails():
    user_id = request.args.get("user_id") or abort(401)
    # Select emails for this user that were sent/drafted and that have an original_content field
    try:
        emails = (
            supabase.table("emails")
                    .select("id, sender_email, subject, original_content, status, sent_at")
                    .eq("user_id", user_id)
                    .in_("status", ["sent","drafted"])   # treat drafted as 'responded' if you want
                    .order("sent_at", desc=True)
                    .execute()
                    .data
            or []
        )
    except Exception:
        app.logger.exception("failed to load responded emails")
        emails = []

    return render_template("partials/responded_emails.html", emails=emails, user_id=user_id)


@app.route("/dashboard/email/<email_id>")
def dashboard_email_view(email_id):
    """Return a small partial showing full original_content — HTMX call for modal."""
    try:
        rec = supabase.table("emails").select("*").eq("id", email_id).single().execute().data
    except Exception:
        rec = None

    if not rec:
        return "<div class='chart-container'>Email not found.</div>", 404

    return render_template("partials/email_modal.html", email=rec)


@app.route("/dashboard/analytics")
def dashboard_analytics():
    user_id = _require_user()
    return render_template("partials/analytics.html", user_id=user_id)

@app.route("/dashboard/users")
def dashboard_users():
    user_id = _require_user()
    users = supabase.table("profiles").select("id, full_name, email").execute().data or []
    return render_template("partials/users.html", users=users)

@app.route("/dashboard/billing")
def dashboard_billing():
    user_id = _require_user()
    return render_template("partials/billing.html", user_id=user_id)

@app.route("/dashboard/settings", methods=["GET", "POST"])
def dashboard_settings():
    user_id = _require_user()

    # ─── Handle Profile POST ────────────────────────────────
    if request.method == "POST":
        section = request.form.get("section")
        if section == "profile":
            new_display_name = request.form.get("display_name", "").strip()
            new_signature = request.form.get("signature", "").strip()
            supabase.table("profiles").update({
                "display_name": new_display_name,
                "signature": new_signature
            }).eq("id", user_id).execute()

    # ─── Fetch profile & flags ──────────────────────────────
    profile_resp = supabase.table("profiles") \
                           .select("display_name, signature, ai_enabled, smtp_email") \
                           .eq("id", user_id) \
                           .single() \
                           .execute()
    
    profile = profile_resp.data or {
        "display_name": "",
        "signature": "",
        "ai_enabled": False,
        "smtp_email": None
    }

    # ▶ Determine Gmail connection status
    gmail_connected = False
    show_reconnect = False
    
    try:
        toks = supabase.table("gmail_tokens") \
                       .select("credentials") \
                       .eq("user_id", user_id) \
                       .single() \
                       .execute().data
        if toks:
            gmail_connected = True
            creds_payload = toks["credentials"]
            creds = Credentials(
                token=creds_payload["token"],
                refresh_token=creds_payload["refresh_token"],
                token_uri=creds_payload["token_uri"],
                client_id=creds_payload["client_id"],
                client_secret=creds_payload["client_secret"],
                scopes=creds_payload["scopes"],
            )
            show_reconnect = creds.expired
    except Exception:
        app.logger.warning(f"settings: could not check Gmail token for {user_id}")

    # ▶ Render template
    return render_template(
        "partials/settings.html",
        profile=profile,
        user_id=user_id,
        gmail_connected=gmail_connected,
        show_reconnect=show_reconnect
    )

import json

import json
from urllib.parse import unquote

@app.route('/connect_smtp_form', methods=['GET', 'POST'])
def connect_smtp_form():
    if request.method == 'POST':
        # Handle POST request (if needed)
        pass
    
    # Handle GET request
    user_id = request.args.get('user_id')
    
    # Get and decode the email parameter
    email_param = request.args.get('email', '')
    email = unquote(email_param) if email_param else ''
    
    # Initialize with default values
    smtp_host = "smtp.gmail.com"
    imap_host = "imap.gmail.com"
    smtp_port = 587
    imap_port = 993
    
    # Try to get detected settings from the request
    settings_param = request.args.get('settings')
    print(f"Raw settings parameter: {settings_param}")  # Debug
    
    if settings_param:
        try:
            # URL decode the settings parameter first
            decoded_settings = unquote(settings_param)
            print(f"Decoded settings: {decoded_settings}")  # Debug
            
            settings = json.loads(decoded_settings)
            smtp_host = settings.get('smtp_host', smtp_host)
            imap_host = settings.get('imap_host', imap_host)
            smtp_port = settings.get('smtp_port', smtp_port)
            imap_port = settings.get('imap_port', imap_port)
            print(f"Using detected settings: {settings}")  # For debugging
        except (json.JSONDecodeError, TypeError) as e:
            print(f"Error parsing settings: {e}")  # For debugging
            # If JSON parsing fails, fall back to defaults
            pass
    
    print(f"Final values - Email: {email}, SMTP: {smtp_host}:{smtp_port}, IMAP: {imap_host}:{imap_port}")  # For debugging
    
    return render_template('partials/connect_smtp_form.html', 
                         user_id=user_id, 
                         email=email,
                         smtp_host=smtp_host,
                         imap_host=imap_host,
                         smtp_port=smtp_port,
                         imap_port=imap_port)

def disconnect_smtp():
    user_id = request.form.get("user_id")
    supabase.table("profiles").update({
        "smtp_email": None,
        "smtp_enc_password": None,
        "smtp_host": None,
        "imap_host": None
    }).eq("id", user_id).execute()
    return redirect(f"/dashboard/signin?user_id={user_id}")

@app.route("/dashboard/home")
def dashboard_home():
    user_id = request.args.get("user_id")
    if not user_id:
        return "Missing user_id", 401

    # (Same logic as /dashboard for HTMX partial)
    profile_resp = (
        supabase.table("profiles")
                .select("display_name, ai_enabled, email, generate_leases")
                .eq("id", user_id)
                .single()
                .execute()
    )
    if profile_resp.data is None:
        return "Profile query error", 500

    profile         = profile_resp.data
    full_name       = profile.get("display_name", "")
    ai_enabled      = profile.get("ai_enabled", True)
    generate_leases = profile.get("generate_leases", False)

    today     = date.today().isoformat()
    sent_rows = (
        supabase.table("emails")
                .select("sent_at")
                .eq("user_id", user_id)
                .eq("status", "sent")
                .execute()
                .data
        or []
    )
    emails_sent_today = sum(1 for e in sent_rows if e.get("sent_at", "").startswith(today))
    time_saved        = emails_sent_today * 5.5

    token_rows = (
        supabase.table("gmail_tokens")
                .select("credentials")
                .eq("user_id", user_id)
                .execute()
                .data
        or []
    )
    show_reconnect = True
    if token_rows:
        creds_data = token_rows[0]["credentials"]
        try:
            creds = Credentials(
                token=creds_data["token"],
                refresh_token=creds_data["refresh_token"],
                token_uri=creds_data["token_uri"],
                client_id=creds_data["client_id"],
                client_secret=creds_data["client_secret"],
                scopes=creds_data["scopes"],
            )
            show_reconnect = creds.expired
        except Exception:
            pass
    # 4) Count "kits generated" for this user
# (Assuming you flag each transaction row with kit_generated=True)
    kit_rows = (
        supabase
        .table("transactions")
        .select("id")
        .eq("user_id", user_id)
        .eq("kit_generated", True)
        .execute()
        .data
        or []
    )
    kits_generated = len(kit_rows)

    # 5) Compute extra estimated time saved
    # e.g. you save ~15 minutes per generated kit
    PER_KIT_SAVE_MINUTES = 15
    estimated_saved = kits_generated * PER_KIT_SAVE_MINUTES
 

    return render_template(
        "partials/home.html",
        name=full_name,
        user_id=user_id,
        emails_sent=emails_sent_today,
        time_saved=time_saved,
        estimated_saved=estimated_saved,  # new computed value
        kits_generated=kits_generated,      # new computed value
        ai_enabled=ai_enabled,
        show_reconnect=show_reconnect,
        generate_leases=generate_leases,
    )
#----------------------------------------------------------------------
@app.route("/reconnect_gmail")
def reconnect_gmail():
    """Handles both initial connection and reconnection to Gmail"""
    user_id = request.args.get("user_id")
    if not user_id:
        return "Missing user ID", 400

    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": os.environ["GOOGLE_CLIENT_ID"],
 
                "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [os.environ["REDIRECT_URI"]]
            }
        },
        scopes=[
            "https://www.googleapis.com/auth/gmail.send",
            "https://www.googleapis.com/auth/gmail.readonly",
            "https://www.googleapis.com/auth/userinfo.email",
            "https://www.googleapis.com/auth/gmail.compose",
            "openid"
        ]
    )
    flow.redirect_uri = os.environ["REDIRECT_URI"]
    authorization_url, _ = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent",
        state=user_id
    )
    return redirect(authorization_url)
  
@app.route("/connect-smtp", methods=["POST"])
def route_connect_smtp():
    try:
        # Get form data
        user_id = request.form.get("user_id")
        smtp_email = request.form.get("smtp_email")
        smtp_password = request.form.get("smtp_password")
        smtp_host = request.form.get("smtp_host")
        imap_host = request.form.get("imap_host")

        # Debug - log all received form data
        app.logger.info(f"Received form data: {dict(request.form)}")

        # Validate required fields
        missing = []
        if not user_id: missing.append("user_id")
        if not smtp_email: missing.append("smtp_email")
        if not smtp_password: missing.append("smtp_password")
        if not smtp_host: missing.append("smtp_host")
        if not imap_host: missing.append("imap_host")

        if missing:
            return jsonify({
                "status": "error",
                "message": f"Missing fields: {', '.join(missing)}"
            }), 400

        # Encrypt the SMTP password
        token = fernet.encrypt(smtp_password.encode()).decode()

        # Upsert into Supabase
        resp = supabase.table("profiles").upsert({
            "id": user_id,
            "smtp_email": smtp_email,
            "smtp_enc_password": token,
            "smtp_host": smtp_host,
            "imap_host": imap_host
        }, on_conflict="id").execute()

        # Check for empty or failed response
        if not resp.data:
            app.logger.error(f"Supabase upsert failed: {resp}")
            return jsonify({
                "status": "error",
                "message": "Failed to save credentials to database"
            }), 500

        # --- 5) On success, send HX-Redirect so HTMX navigates for us ---
        hxr = make_response("", 204)
        hxr.headers["HX-Redirect"] = url_for("complete_profile", user_id=user_id)
        return hxr

    except Exception as e:
        app.logger.error("connect-smtp error", exc_info=True)
        return jsonify({
            "status": "error", 
            "message": "Internal server error"
        }), 500



#------------------------------------------ 


#------------------------------------------
# Update your send_email function to require valid connection
@app.route("/send", methods=["POST"])
def send_email():
    data = request.get_json()
    user_id = data["user_id"]
    
    # Check email connection before proceeding
    require_valid_email_connection(user_id)
    
    # Rest of your send logic...
    to = data["to"]
    subject = data["subject"]
    body = data["body"]

    smtp_email, app_password = get_smtp_creds(user_id)
    if smtp_email and app_password:
        # Use SMTP fallback
        send_email_smtp(smtp_email, app_password, to, subject, body)
        return jsonify({"method": "smtp", "status": "sent"}), 200
        return jsonify({"method": "gmail", "messages": []}), 200
    # else: your existing Gmail API flow
    return send_via_gmail_api(data)

# Update your fetch_mail function to require valid connection
@app.route("/fetch", methods=["GET"])
def fetch_mail():
    user_id = request.args.get("user_id")
    
    # Check email connection before proceeding
    require_valid_email_connection(user_id)
    
    smtp_email, app_password = get_smtp_creds(user_id)
    if smtp_email and app_password:
        messages = fetch_emails_imap(smtp_email, app_password)
        return jsonify({"method": "imap", "messages": messages}), 200
        return jsonify({"method": "gmail", "messages": []}), 200
    # else: your existing Gmail-API‐based fetch
    return fetch_via_gmail_api(user_id)

# Add this route to update the connection status in the database
@app.route("/update_connection_status", methods=["POST"])
def update_connection_status():
    user_id = request.form.get("user_id")
    status = request.form.get("status")  # "valid" or "invalid"
    
    if not user_id or not status:
        return jsonify({"status": "error", "message": "Missing parameters"}), 400
    
    # Update the connection status in the database
    supabase.table("profiles").update({
        "email_connection_status": status,
        "connection_checked_at": datetime.now(timezone.utc).isoformat()
    }).eq("id", user_id).execute()
    
    return jsonify({"status": "success"})
#-----------------------------------------------------------------------
@app.route("/connect_gmail")
def connect_gmail():
    """
    Initiates Gmail OAuth flow.
    """
    flow = Flow.from_client_config(
        {
            "web": {
                "client_id": os.environ["GOOGLE_CLIENT_ID"],
                "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
                "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                "token_uri": "https://oauth2.googleapis.com/token",
                "redirect_uris": [os.environ["REDIRECT_URI"]]
            }
        },
        scopes=[
            "https://www.googleapis.com/auth/gmail.send",
            "https://www.googleapis.com/auth/gmail.readonly",
            "https://www.googleapis.com/auth/userinfo.email",
            "https://www.googleapis.com/auth/gmail.compose",
            "openid"
        ]
    )
    flow.redirect_uri = os.environ["REDIRECT_URI"]
    authorization_url, _ = flow.authorization_url(
        access_type="offline",
        include_granted_scopes="true",
        prompt="consent"
    )
    return redirect(authorization_url)

@app.route("/oauth2callback")
def oauth2callback():
    """Handles OAuth2 callback from Google"""
    try:
        # Extract state parameter containing user_id
        user_id = request.args.get("state")
        if not user_id:
            app.logger.error("OAuth2 callback missing state parameter")
            return "<h1>Authentication Failed</h1><p>Missing state parameter</p>", 400
        
        flow = Flow.from_client_config(
            {
                "web": {
                    "client_id": os.environ["GOOGLE_CLIENT_ID"],
                    "client_secret": os.environ["GOOGLE_CLIENT_SECRET"],
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [os.environ["REDIRECT_URI"]]
                }
            },
            scopes=[
                "https://www.googleapis.com/auth/gmail.send",
                "https://www.googleapis.com/auth/gmail.readonly",
                "https://www.googleapis.com/auth/userinfo.email",
                "https://www.googleapis.com/auth/gmail.compose",
                "openid"
            ],
            state=user_id
        )
        flow.redirect_uri = os.environ["REDIRECT_URI"]
        flow.fetch_token(authorization_response=request.url)
        credentials = flow.credentials

        # Verify ID token
        id_info = id_token.verify_oauth2_token(
            credentials.id_token,
            grequests.Request(),
            os.environ["GOOGLE_CLIENT_ID"]
        )
        email = id_info.get("email")
        if not email:
            raise ValueError("No email found in Google ID token")

        # Upsert gmail tokens
        creds_payload = {
            "user_id": user_id,
            "user_email": email,
            "credentials": {
                "token": credentials.token,
                "refresh_token": credentials.refresh_token,
                "token_uri": credentials.token_uri,
                "client_id": credentials.client_id,
                "client_secret": credentials.client_secret,
                "scopes": credentials.scopes
            }
        }
        supabase.table("gmail_tokens").upsert(creds_payload).execute()

        # Update user profile
        full_name = id_info.get("name") or email.split("@")[0]
        supabase.table("profiles").update({
            "email": email,
            "full_name": full_name,
            "ai_enabled": True
        }).eq("id", user_id).execute()

        return redirect(f"/dashboard?user_id={user_id}")

    except Exception as e:
        app.logger.error(f"OAuth2 Callback Error: {str(e)}", exc_info=True)
        return f"<h1>Authentication Failed</h1><p>{str(e)}</p>", 500
      
@app.route("/complete_profile", methods=["GET", "POST"])
def complete_profile():
    user_id = request.args.get("user_id")
    if not user_id:
        return "Missing user_id", 401

    if request.method == "POST":
        display_name = request.form.get("display_name", "").strip()
        signature    = request.form.get("signature", "").strip()

        supabase.table("profiles") \
            .update({"display_name": display_name,
                     "signature": signature}) \
            .eq("id", user_id) \
            .execute()

        return redirect(f"/dashboard?user_id={user_id}")

    return render_template("complete_profile.html", user_id=user_id)

@app.route("/disconnect_gmail", methods=["POST"])
def disconnect_gmail():
    user_id = request.form.get("user_id")
    supabase.table("gmail_tokens").delete().eq("user_id", user_id).execute()
    return redirect(f"/dashboard?user_id={user_id}")

def _require_user():
    uid = request.args.get("user_id") or request.form.get("user_id")
    if not uid:
        abort(401, "Missing user_id")
    return uid

@app.route("/new_lease", methods=["GET"])
def new_lease_form():
    user_id = _require_user()
    return render_template("new_lease.html", user_id=user_id)

@app.route("/new_lease", methods=["POST"])
def new_lease_submit():
    user_id = _require_user()

    data = {
        "property_name":    request.form["propertyName"],
        "property_type":    request.form["propertyType"],
        "address":          request.form["address"],
        "suite":            request.form.get("suite",""),
        "square_feet":      request.form["squareFeet"],
        "tenant_name":      request.form["tenantName"],
        "tenant_type":      request.form["tenantType"],
        "lease_type":       request.form["leaseType"],
        "lease_term":       request.form["leaseTerm"],
        "start_date":       request.form["startDate"],
        "end_date":         request.form["endDate"],
        "base_rent":        request.form["baseRent"],
        "annual_increase":  request.form.get("annualIncrease",""),
        "security_deposit": request.form.get("securityDeposit",""),
        "parking_spaces":   request.form.get("parkingSpaces",""),
        "parking_fee":      request.form.get("parkingFee",""),
        "additional_terms": request.form.get("additionalTerms",""),
        "tenant_improvements": "Yes" if request.form.get("tenantImprovements") else "No",
        "renewal_option":      "Yes" if request.form.get("renewalOption") else "No",
        "exclusive_use":       "Yes" if request.form.get("exclusiveUse") else "No",
    }

    html_body = f"""
    <html><body>
      <h2>Lease Agreement</h2>
      <p><strong>Property:</strong> {data['property_name']} ({data['property_type'].title()})<br>
      <strong>Address:</strong> {data['address']} Suite {data['suite']}<br>
      <strong>Size:</strong> {data['square_feet']} sqft</p>

      <h3>Tenant</h3>
      <p>{data['tenant_name']} ({data['tenant_type'].title()})</p>

      <h3>Terms</h3>
      <p><strong>Type:</strong> {data['lease_type'].replace('-', '').title()}<br>
      <strong>Term:</strong> {data['lease_term']} months<br>
      <strong>Dates:</strong> {data['start_date']} → {data['end_date']}</p>

      <h3>Financials</h3>
      <p><strong>Base Rent:</strong> ${data['base_rent']} per sqft/yr<br>
      <strong>Annual Increase:</strong> {data['annual_increase']}%<br>
      <strong>Security Deposit:</strong> ${data['security_deposit']}<br>
      <strong>Parking:</strong> {data['parking_spaces']} spaces @ ${data['parking_fee']}/mo</p>

      <h3>Additional Terms</h3>
      <p>{data['additional_terms']}</p>
      <ul>
        <li>Tenant Improvements: {data['tenant_improvements']}</li>
        <li>Renewal Option: {data['renewal_option']}</li>
        <li>Exclusive Use Clause: {data['exclusive_use']}</li>
      </ul>
    </body></html>
    """

    tok = (supabase.table("gmail_tokens")
                .select("credentials")
                .eq("user_id", user_id)
                .limit(1)
                .execute()
                .data) or []
    if not tok:
        abort(400, "No Gmail token; reconnect Gmail first.")

    cd = tok[0]["credentials"]
    creds = Credentials(
        token=cd["token"],
        refresh_token=cd["refresh_token"],
        token_uri=cd["token_uri"],
        client_id=cd["client_id"],
        client_secret=cd["client_secret"],
        scopes=cd["scopes"],
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(GoogleRequest())

    service = build("gmail", "v1", credentials=creds, cache_discovery=False)

    mime = MIMEText(html_body, "html")
    mime["To"]      = ""
    mime["Subject"] = f"Draft Lease: {data['property_name']} → {data['tenant_name']}"
    raw = base64.urlsafe_b64encode(mime.as_bytes()).decode()
    draft = {"message": {"raw": raw}}
    created = service.users().drafts().create(userId="me", body=draft).execute()

    app.logger.info(f"Gmail Draft {created['id']} created for user {user_id}")

    return redirect(f"/dashboard?user_id={user_id}")

@app.route("/admin")
def admin():
    return render_template("admin.html")

@app.route("/api/admin/users")
def api_admin_users():
    users = supabase.table("profiles").select("*").execute().data or []
    today = date.today().isoformat()
    results = []
    for user in users:
        sent = supabase.table("emails") \
            .select("sent_at") \
            .eq("user_id", user["id"]) \
            .eq("status", "sent") \
            .execute().data or []
        count = len([e for e in sent if e["sent_at"] and e["sent_at"].startswith(today)])
        results.append({
            "id": user["id"],
            "name": user["full_name"],
            "email": user["email"],
            "enabled": user.get("ai_enabled", True),
            "emails_today": count
        })
    return jsonify(results)

@app.route("/api/admin/toggle_status", methods=["POST"])
def api_toggle_status():
    user_id = request.json.get("user_id")
    enable = request.json.get("enable", True)
    supabase.table("profiles").update({"ai_enabled": enable}).eq("id", user_id).execute()
    return jsonify({"success": True})

@app.route("/debug_env")
def debug_env():
    return {
        "GOOGLE_CLIENT_ID": os.environ.get("GOOGLE_CLIENT_ID"),
        "REDIRECT_URI": os.environ.get("REDIRECT_URI"),
        "EDGE_BASE_URL": os.environ.get("EDGE_BASE_URL")
    }

from datetime import datetime

@app.route("/process", methods=["GET"])
@check_rate_limit('emails')
def trigger_process():
    token = request.args.get("token")
    if token != os.environ.get("PROCESS_SECRET_TOKEN"):
        return jsonify({"error": "Unauthorized"}), 401
        
    # Decrement email count
#    ip = request.remote_addr
 #   demo_rate_limits[ip]['emails'] -= 1
    
# ── 0) DAILY RESET CHECK ──
    today_str = date.today().isoformat()
    rl_row = SUPABASE_SERVICE.table("rate_limit_reset") \
        .select("last_reset") \
        .eq("id", "global") \
        .single() \
        .execute().data or {}
    last_date = rl_row.get("last_reset", "")[:10]  # e.g. "2025-07-27"

    if last_date != today_str:
        app.logger.info("🔄 New day detected – clearing emails table")

        # Delete all rows by filtering out a UUID value that never exists
        SUPABASE_SERVICE.table("emails") \
            .delete() \
            .neq("id", "00000000-0000-0000-0000-000000000000") \
            .execute()

        # Update the reset timestamp
        SUPABASE_SERVICE.table("rate_limit_reset") \
            .update({"last_reset": datetime.now(timezone.utc).isoformat()}) \
            .eq("id", "global") \
            .execute()

        
    # ── 0) Build per-user counts of emails already sent today (YYYY‑MM‑DD) ──
    today_iso = datetime.utcnow().date().isoformat()
    sent_rows = (
        supabase.table("emails")
                .select("user_id, sent_at")
                .eq("status", "sent")
                .execute()
                .data or []
    )
    emails_sent_today: dict[str,int] = {}
    for r in sent_rows:
        sent_at = r.get("sent_at","")
        if sent_at.startswith(today_iso):
            uid = r["user_id"]
            emails_sent_today[uid] = emails_sent_today.get(uid, 0) + 1

    # ── 1) Fetch the three pre‑send queues ──
    gen  = supabase.table("emails").select("id").eq("status", "processing").execute().data or []
    per  = supabase.table("emails").select("id").eq("status", "ready_to_personalize").execute().data or []
    prop = supabase.table("emails").select("id").eq("status", "awaiting_proposal").execute().data or []

    if not (gen or per or prop):
        app.logger.info("⚡ No emails to process — returning 204")
        return "", 204

    all_processed, sent, drafted, failed = [], [], [], []

    # ── 2) Generate Response ──
    if gen:
        # Check rate limit before making AI calls
        ip = request.remote_addr
        now = datetime.now()
        
        # Reset email limit if it's a new day
        if (now - demo_rate_limits[ip]['emails']['last_reset']).days >= 1:
            demo_rate_limits[ip]['emails']['remaining'] = 20
            demo_rate_limits[ip]['emails']['last_reset'] = now
        
        # Check if we have remaining emails
        if demo_rate_limits[ip]['emails']['remaining'] <= 0:
            app.logger.warning(f"Rate limit exceeded for IP {ip}, skipping AI calls")
            # Mark emails as error due to rate limiting
            ids = [r["id"] for r in gen]
            supabase.table("emails")\
                    .update({"status":"error","error_message":"Rate limit exceeded"})\
                    .in_("id", ids).execute()
        else:
            # Decrement the counter and proceed with AI calls
            demo_rate_limits[ip]['emails']['remaining'] -= 1
            ids = [r["id"] for r in gen]
            if call_edge("/functions/v1/clever-service/generate-response", {"email_ids": ids}):
                all_processed.extend(ids)
            else:
                supabase.table("emails")\
                        .update({"status":"error","error_message":"generate-response failed"})\
                        .in_("id", ids).execute()


    # ── 3) Personalize Template ──
    if per:
        for eid in [r["id"] for r in per]:
            if call_edge("/functions/v1/clever-service/personalize-template", {"email_ids":[eid]}):
                supabase.table("emails").update({"status":"awaiting_proposal"}).eq("id", eid).execute()
                all_processed.append(eid)
            else:
                supabase.table("emails")\
                        .update({"status":"error","error_message":"personalize-template failed"})\
                        .eq("id", eid).execute()

    # ── 4) Generate Proposal → ready_to_send ──
    if prop:
        for eid in [r["id"] for r in prop]:
            if call_edge("/functions/v1/clever-service/generate-proposal", {"email_ids":[eid]}):
                supabase.table("emails").update({"status":"ready_to_send"}).eq("id", eid).execute()
                all_processed.append(eid)
            else:
                supabase.table("emails")\
                        .update({"status":"error","error_message":"generate-proposal failed"})\
                        .eq("id", eid).execute()

    # ── 5) Re‑fetch ready_to_send rows ──
    ready = (
        supabase.table("emails")
                .select("id, user_id, sender_email, processed_content, subject")
                .eq("status", "ready_to_send")
                .execute()
                .data or []
    )

        # ── 6) Send via SMTP fallback or Gmail API, enforcing 20/day cap ──
    for rec in ready:
        em_id     = rec["id"]
        uid       = rec["user_id"]
        to_addr   = rec["sender_email"]
        subject   = rec.get("subject", "Your Email")  # Get the original subject or default

        # 20-email/day limit
        if emails_sent_today.get(uid, 0) >= 20:
            app.logger.info(f"User {uid} reached daily limit, marking {em_id} error")
            supabase.table("emails").update({
                "status": "error",
                "error_message": "Daily email limit reached"
            }).eq("id", em_id).execute()
            failed.append(em_id)
            continue 

        # load personalization flags & build HTML
        lease_flag = supabase.table("profiles") \
                             .select("generate_leases") \
                             .eq("id", uid).single().execute().data.get("generate_leases", False)
        body_html = (rec.get("processed_content") or "").replace("\n", "<br>")
        prof_sig = supabase.table("profiles") \
                           .select("display_name, signature") \
                           .eq("id", uid).single().execute().data or {}
        if prof_sig.get("display_name"):
            body_html = body_html.replace("[Your Name]", prof_sig["display_name"])
        full_html = f"<html><body><p>{body_html}</p>{prof_sig.get('signature','')}</body></html>"

        # 20-email/day limit
        if emails_sent_today.get(uid, 0) >= 20:
            app.logger.info(f"User {uid} reached daily limit, marking {em_id} error")
            supabase.table("emails").update({
                "status": "error",
                "error_message": "Daily email limit reached"
            }).eq("id", em_id).execute()
            failed.append(em_id)
            continue

        # 1) SMTP fallback
        prof = supabase.table("profiles") \
                       .select("smtp_email,smtp_enc_password,smtp_host") \
                       .eq("id", uid).single().execute().data or {}
        if prof.get("smtp_email") and prof.get("smtp_enc_password"):
            smtp_email = prof["smtp_email"]
            smtp_pass  = fernet.decrypt(prof["smtp_enc_password"].encode()).decode()
            smtp_host  = prof.get("smtp_host", "smtp.gmail.com")
            try:
                send_email_smtp(
                    smtp_email,
                    smtp_pass,
                    to_addr,
                    "Lease Agreement Draft" if lease_flag else f"RE: {rec.get('subject', 'Your Email')}",  # Modified subject,
                    full_html,
                    smtp_host=smtp_host
                )
                supabase.table("emails").update({
                    "status":  "sent",
                    "sent_at": datetime.utcnow().isoformat()
                }).eq("id", em_id).execute()
                emails_sent_today[uid] = emails_sent_today.get(uid, 0) + 1
                sent.append(em_id)
                app.logger.info(f"SMTP send succeeded for email {em_id} (user {uid})")
            except Exception as e:
                app.logger.error(f"SMTP send failed for email {em_id} (user {uid})", exc_info=True)
                supabase.table("emails").update({
                    "status":        "error",
                    "error_message": str(e)
                }).eq("id", em_id).execute()
                failed.append(em_id)
            continue  # next `rec`

        # 2) Gmail API fallback
        try:
            tok = supabase.table("gmail_tokens") \
                          .select("credentials") \
                          .eq("user_id", uid).single().execute().data
            if not tok:
                raise ValueError("No Gmail token found")

            cd = tok["credentials"]
            creds = Credentials(
                token=cd["token"],
                refresh_token=cd["refresh_token"],
                token_uri=cd["token_uri"],
                client_id=cd["client_id"],
                client_secret=cd["client_secret"],
                scopes=cd["scopes"],
            )
            if creds.expired and creds.refresh_token:
                creds.refresh(GoogleRequest())

            svc = build("gmail", "v1", credentials=creds, cache_discovery=False)
            msg = MIMEText(full_html, "html")
            msg["to"]      = to_addr
            msg["from"]    = "me"
            msg["subject"] = "Lease Agreement Draft" if lease_flag else f"RE: {rec.get('subject', 'Your Email')}"  # Modified subject
            raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()

            if lease_flag:
                svc.users().drafts().create(userId="me", body={"message": {"raw": raw}}).execute()
                status_to = "drafted"
                drafted.append(em_id)
            else:
                svc.users().messages().send(userId="me", body={"raw": raw}).execute()
                status_to = "sent"
                sent.append(em_id)

            supabase.table("emails").update({
                "status":  status_to,
                "sent_at": datetime.utcnow().isoformat()
            }).eq("id", em_id).execute()
            emails_sent_today[uid] = emails_sent_today.get(uid, 0) + 1
            app.logger.info(f"Gmail API send succeeded for email {em_id} (user {uid})")

        except Exception as e:
            app.logger.error(f"Gmail API send failed for email {em_id} (user {uid})", exc_info=True)
            supabase.table("emails").update({
                "status":        "error",
                "error_message": str(e)
            }).eq("id", em_id).execute()
            failed.append(em_id)

       

  
    # ── Summary response ──
    summary = {
        "processed": all_processed,
        "sent":      sent,
        "drafted":   drafted,
        "failed":    failed
    }
    return jsonify(summary), 200

#---------------------------------------------------------------------------------------------------------------------------



#-----------------------------------------------------------------------------------------------------

@app.route("/transaction/<txn_id>/ready", methods=["POST"])
def mark_ready(txn_id):
    supabase.table("transactions").update({"ready_for_kit": True}).eq("id", txn_id).execute()
    return "", 204

@app.route("/autopilot/batch", methods=["POST"])
@check_rate_limit('kits')
def batch_autopilot():
    # Decrement kit count
#    ip = request.remote_addr
 #   demo_rate_limits[ip]['kits'] -= 1
    
    txns = supabase.table("transactions").select("*").eq("ready_for_kit", True).eq("kit_generated", False).execute().data or []
    results = []
    for t in txns:
        payload = {
          "transaction_type": t["transaction_type"],
          "data": {
            "id": t["id"],
            "buyer": t["buyer"],
            "seller": t["seller"],
            "date": t["date"],
            "purchase_price": t["purchase_price"],
            "closing_date": t.get("closing_date"),
            "closing_location": t.get("closing_location")
          }
        }
        resp = requests.post(f"{os.environ.get('BASE_URL')}/autopilot/trigger", json=payload)
        results.append({"id": t["id"], "status": resp.status_code})
        if resp.ok:
            supabase.table("transactions").update({"kit_generated": True}).eq("id", t["id"]).execute()
    return jsonify(results), 200

@app.route("/dashboard/autopilot")
def dashboard_autopilot():
    user_id = request.args.get("user_id") or abort(401)
    txn_id  = request.args.get("txn_id")
    transactions = supabase.table("transactions").select("*").eq("user_id", user_id).execute().data or []
    current_txn = None
    if txn_id:
        resp = supabase.table("transactions").select("*").eq("id", txn_id).execute()
        current_txn = resp.data[0] if resp.data else None
    return render_template("partials/autopilot.html", user_id=user_id, transactions=transactions, current_transaction=current_txn)

@app.route("/transactions/new", methods=["POST"])
def create_transaction():
    import uuid
    import traceback

    user_id = request.args.get("user_id") or request.form.get("user_id")
    if not user_id:
        return jsonify({"status": "error", "message": "Missing user_id"}), 401

    new_id = str(uuid.uuid4())

    # 🔐 Validate required fields (lowercase unified names)
    required = ["buyer_name", "seller_name", "property_address", "agreement_date"]
    missing = [f for f in required if not request.form.get(f)]
    if missing:
        app.logger.warning(f"⚠️ Missing required fields: {missing}")
        return jsonify({
            "status": "error",
            "message": f"Missing required fields: {', '.join(missing)}"
        }), 400

    # ✅ All accepted lowercase fields from gamified form
    accepted_fields = [
        "transaction_type", "property_address", "city", "state", "name_of_property",
        "description_of_property", "square_feet", "legal_description",
        "apartment_address", "premises_description",

        "buyer_name", "buyer_address", "seller_name", "seller_address", "agency_name",

        "purchase_price", "deposit_amount", "agreement_date", "broker_name",
        "commission_amount", "brokerage_fee", "broker_payday",

        "closing_date", "occupy_property_date", "mortgage_amount", "mortgage_years",
        "interest_rate", "inspection_days", "possession_date",

        "rent_type", "agreed_rent", "maintenance_terms",

        "landlord_phone", "tenant_phone", "landlord_email", "tenant_email",

        "structure_age", "location", "county", "additional_explanations",

        "buyer_signature", "seller_signature", "time"
    ]

    # Build the payload, turning empty strings into None
    payload = {"id": new_id, "user_id": user_id}
    for field in accepted_fields:
        val = request.form.get(field)
        payload[field] = None if val == "" else val

    try:
        app.logger.info(f"🚀 Inserting transaction with ID {new_id}")
        app.logger.debug(f"Payload: {payload}")
        resp = supabase.table("transactions").insert(payload).execute()
        inserted = resp.data[0]
    except Exception as e:
        app.logger.error("❌ Transaction insert failed")
        app.logger.error(traceback.format_exc())
        return jsonify({
            "status": "error",
            "message": f"Insertion failed: {str(e)}"
        }), 500

    # ✅ Success response with htmx trigger
    feedback = (
        f'<div class="alert alert-success">🎉 Transaction <strong>{inserted["id"]}</strong> created.</div>'
        + '<script>htmx.trigger(document.querySelector(\'[hx-get*="/dashboard/autopilot"]\'), "click")</script>'
    )
    return feedback, 200

# Add this to your main app file (e.g., app.py)
# Add these imports at the top of your app.py
import re
import dns.resolver

# Add this route to your app.py


def extract_domain(email):
    """Extract domain from email address"""
    pattern = r'@([\w\.-]+)'
    match = re.search(pattern, email)
    if match:
        return match.group(1).lower()
    return None


@app.route("/check_smtp_status")
def check_smtp_status():
    user_id = request.args.get("user_id")
    if not user_id:
        return jsonify({"status": "error", "message": "Missing user_id"}), 400
    
    # Get SMTP credentials
    smtp_email, app_password = get_smtp_creds(user_id)
    if not smtp_email or not app_password:
        return jsonify({"status": "invalid", "message": "No SMTP credentials found"})
    
    # Get SMTP server details from profile
    resp = supabase.from_("profiles").select("smtp_host, imap_host").eq("id", user_id).single().execute()
    if not resp.data:
        return jsonify({"status": "error", "message": "Could not retrieve server details"}), 500
    
    server_details = resp.data
    smtp_host = server_details.get("smtp_host", "smtp.gmail.com")
    smtp_port =  587
    imap_host = server_details.get("imap_host", "imap.gmail.com")
    imap_port = 993
    
    # Test SMTP connection
    smtp_working = False
    imap_working = False
    
    try:
        # Test SMTP
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_email, app_password)
            smtp_working = True
    except Exception as e:
        app.logger.error(f"SMTP test failed for {user_id}: {str(e)}")
    
    try:
        # Test IMAP
        with imaplib.IMAP4_SSL(imap_host, imap_port) as server:
            server.login(smtp_email, app_password)
            imap_working = True
    except Exception as e:
        app.logger.error(f"IMAP test failed for {user_id}: {str(e)}")
    
    if smtp_working and imap_working:
        return jsonify({"status": "valid", "message": "SMTP and IMAP connections successful"})
    elif smtp_working:
        return jsonify({"status": "partial", "message": "SMTP working but IMAP failed"})
    else:
        return jsonify({"status": "invalid", "message": "Both SMTP and IMAP failed"})
def check_smtp_status_alias():
    return check_email_connection()

#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

import csv
from io import TextIOWrapper
from openpyxl import load_workbook

@app.route("/import_leads", methods=["GET", "POST"])
@check_rate_limit('leads')
def import_leads():
    user_id = _require_user()
    
    if request.method == "GET":
        return render_template("import_leads.html", user_id=user_id)
    
    # Handle POST request
    try:
        # Debug logging
        app.logger.info(f"Import leads request received: {request.files}")
        
        if 'file' not in request.files:
            app.logger.error("No file in request")
            return jsonify({"error": "No file uploaded"}), 400
        
        file = request.files['file']
        if file.filename == '':
            app.logger.error("Empty filename")
            return jsonify({"error": "No file selected"}), 400
        
        # Check file extension
        if file.filename.endswith('.csv'):
            # Process CSV file
            csv_file = TextIOWrapper(file, encoding='utf-8')
            reader = csv.DictReader(csv_file)
            rows = list(reader)
            app.logger.info(f"CSV columns: {reader.fieldnames}")
        elif file.filename.endswith(('.xlsx', '.xls')):
            # Process Excel file
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers
            headers = [cell.value for cell in ws[1] if cell.value]
            
            # Get data rows
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                # Remove None values
                rows.append({k: v for k, v in row_data.items() if v is not None})
            
            app.logger.info(f"Excel columns: {headers}")
        else:
            app.logger.error(f"Invalid file type: {file.filename}")
            return jsonify({"error": "Invalid file type. Please upload CSV or Excel."}), 400
        
        # Check if we have any rows
        if not rows:
            app.logger.error("No data rows found in file")
            return jsonify({"error": "No data found in file"}), 400
        
        # Check required columns (more flexible approach)
        required_columns = ['name', 'Last name', 'Recipient', 'city', 'brokerage', 'service', 'Email Sent']
        available_columns = list(rows[0].keys())
        
        app.logger.info(f"Available columns: {available_columns}")
        app.logger.info(f"Required columns: {required_columns}")
        
        missing_columns = [col for col in required_columns if col not in available_columns]
        if missing_columns:
            app.logger.error(f"Missing columns: {missing_columns}")
            return jsonify({
                "error": f"Missing required columns: {', '.join(missing_columns)}. "
                        f"Available columns: {', '.join(available_columns)}"
            }), 400
        
        # Process each row
        success_count = 0
        error_count = 0
        imported_leads = []
        
        for i, row in enumerate(rows):
            try:
                # Parse email_sent date
                email_sent_str = row.get('Email Sent', '')
                email_sent = None
                
                try:
                    if email_sent_str:
                        # Try different date formats
                        if isinstance(email_sent_str, str):
                            try:
                                email_sent = datetime.strptime(email_sent_str, '%Y-%m-%d %H:%M:%S')
                            except ValueError:
                                try:
                                    email_sent = datetime.strptime(email_sent_str, '%Y-%m-%d')
                                except ValueError:
                                    # Try to parse Excel serial date numbers
                                    try:
                                        if isinstance(email_sent_str, (int, float)):
                                            email_sent = datetime(1899, 12, 30) + timedelta(days=email_sent_str)
                                        else:
                                            email_sent = datetime.utcnow()
                                    except:
                                        email_sent = datetime.utcnow()
                        else:
                            # Assume it's already a datetime object
                            email_sent = email_sent_str
                    else:
                        email_sent = datetime.utcnow()
                except Exception as e:
                    app.logger.warning(f"Error parsing date {email_sent_str}: {e}")
                    email_sent = datetime.utcnow()
                
                # Prepare lead data
                lead_data = {
                    'user_id': user_id,
                    'first_name': str(row.get('name', '')).strip(),
                    'last_name': str(row.get('Last name', '')).strip(),
                    'email': str(row.get('Recipient', '')).strip().lower(),
                    'city': str(row.get('city', '')).strip(),
                    'brokerage': str(row.get('brokerage', '')).strip(),
                    'service': str(row.get('service', '')).strip(),
                    'status': 'new',
                    'email_sent': email_sent.isoformat() if hasattr(email_sent, 'isoformat') else email_sent,
                    'created_at': datetime.utcnow().isoformat()
                }
                
                # Validate required fields
                if not lead_data['email'] or '@' not in lead_data['email']:
                    app.logger.warning(f"Row {i+1}: Invalid email address '{lead_data['email']}'")
                    error_count += 1
                    continue
                
                if not lead_data['first_name'] and not lead_data['last_name']:
                    app.logger.warning(f"Row {i+1}: Missing both first and last name")
                    error_count += 1
                    continue
                
                # Insert lead
                response = supabase.table('leads').insert(lead_data).execute()
                
                if response.data:
                    success_count += 1
                    lead_id = response.data[0]['id']
                    imported_leads.append(lead_id)
                    
                    # Schedule follow-ups
                    try:
                        # Send immediate follow-up (step 0)
                        follow_up_content = generate_follow_up_content(lead_id, 0)
                        
                        if follow_up_content:
                            # Get lead details
                            lead = supabase.table('leads').select('*').eq('id', lead_id).single().execute().data
                            
                            # Get user's email credentials
                            smtp_email, app_password = get_smtp_creds(user_id)
                            if smtp_email and app_password:
                                # Get SMTP server details
                                prof_resp = supabase.from_("profiles").select("smtp_host").eq("id", user_id).single().execute()
                                smtp_host = prof_resp.data.get("smtp_host", "smtp.gmail.com") if prof_resp.data else "smtp.gmail.com"
                                
                                # Send the email
                                send_email_smtp(
                                    smtp_email,
                                    app_password,
                                    lead['email'],
                                    "Follow-up from your inquiry",
                                    follow_up_content,
                                    smtp_host=smtp_host
                                )
                                
                                # Create follow-up record
                                follow_up_data = {
                                    'lead_id': lead_id,
                                    'sequence_step': 0,
                                    'generated_content': follow_up_content,
                                    'status': 'sent',
                                    'sent_at': datetime.utcnow().isoformat()
                                }
                                supabase.table('lead_follow_ups').insert(follow_up_data).execute()
                    
                    except Exception as e:
                        app.logger.error(f"Error sending immediate follow-up for lead {lead_id}: {str(e)}")
                    
                    # Schedule the rest of the follow-up sequence
                    for step, seq in enumerate(FOLLOW_UP_SEQUENCE[1:], start=1):
                        scheduled_at = email_sent + timedelta(days=seq['delay_days'])
                        follow_up_data = {
                            'lead_id': lead_id,
                            'sequence_step': step,
                            'scheduled_at': scheduled_at.isoformat(),
                            'status': 'pending'
                        }
                        supabase.table('lead_follow_ups').insert(follow_up_data).execute()
                
                else:
                    error_count += 1
                    app.logger.error(f"Failed to insert lead: {response}")
            
            except Exception as e:
                error_count += 1
                app.logger.error(f"Error processing row {i+1}: {e}", exc_info=True)
        
        # Log summary
        app.logger.info(f"Import completed: {success_count} succeeded, {error_count} failed")
        
        return jsonify({
            "message": f"Leads imported successfully. {success_count} succeeded, {error_count} failed.",
            "imported_count": success_count,
            "failed_count": error_count
        }), 200
    
    except Exception as e:
        app.logger.error(f"Error importing leads: {str(e)}", exc_info=True)
        return jsonify({"error": f"Failed to import leads: {str(e)}"}), 500
#------------------------------------------------------------------------------------------------------------------
def generate_follow_up_content(lead_id, sequence_step):
    """Generate follow-up content using AI with context of previous communications"""
    try:
        app.logger.info(f"Starting follow-up generation for lead {lead_id}, step {sequence_step}")
        
        # Get lead details
        lead_resp = supabase.table("leads").select("*").eq("id", lead_id).single().execute()
        if not lead_resp.data:
            app.logger.error(f"Lead {lead_id} not found")
            return None
            
        lead = lead_resp.data
        app.logger.info(f"Found lead: {lead['email']}")
        
        # Get previous emails from emails table
        previous_emails = supabase.table("emails") \
            .select("subject, original_content, processed_content, sent_at") \
            .eq("sender_email", lead["email"]) \
            .order("sent_at", desc=True) \
            .limit(5) \
            .execute().data or []
        
        # Get previous follow-ups from lead_follow_ups table
        previous_follow_ups = supabase.table("lead_follow_ups") \
            .select("generated_content, sent_at, sequence_step") \
            .eq("lead_id", lead_id) \
            .eq("status", "sent") \
            .lt("sequence_step", sequence_step) \
            .order("sent_at", desc=True) \
            .execute().data or []
        
        app.logger.info(f"Found {len(previous_emails)} previous emails and {len(previous_follow_ups)} previous follow-ups")
        
        # Build context for AI
        context = f"""
        Lead: {lead['first_name']} {lead['last_name']}
        Company: {lead['brokerage']}
        Service: {lead['service']}
        Location: {lead['city']}
        
        Previous communications:
        """
        
        # Add emails from emails table
        for i, email in enumerate(previous_emails):
            context += f"\nEmail {i+1} ({email.get('sent_at', '')}):\n"
            context += f"Subject: {email.get('subject', 'No subject')}\n"
            content = email.get('original_content') or email.get('processed_content', '')
            context += f"Content: {content[:200]}...\n" if len(content) > 200 else f"Content: {content}\n"
        
        # Add follow-ups from lead_follow_ups table
        for i, follow_up in enumerate(previous_follow_ups, start=len(previous_emails)+1):
            context += f"\nFollow-up {i} (Day {FOLLOW_UP_SEQUENCE[follow_up['sequence_step']]['delay_days']}, {follow_up.get('sent_at', '')}):\n"
            content = follow_up.get('generated_content', '')
            context += f"Content: {content[:200]}...\n" if len(content) > 200 else f"Content: {content}\n"
        
        if not previous_emails and not previous_follow_ups:
            context += "\nNo previous communications found. This is the first contact.\n"
        
        context += f"\n\nWrite a friendly, professional follow-up email for day {FOLLOW_UP_SEQUENCE[sequence_step]['delay_days']}."
        context += " Reference previous communications if relevant. Keep it concise and focused on providing value."
        
        app.logger.info(f"Built context for AI: {context[:500]}...")
        
        # Call your AI API
        payload = {
            "context": context,
            "type": "follow_up",
            "sequence_step": sequence_step,
            "lead_id": lead_id
        }
        
        app.logger.info(f"Calling edge function with payload: {payload}")
        
        # Use your existing Edge Function call pattern
        # Modify call_edge to return the response content instead of just success/failure
        response = call_edge("/functions/v1/generate-follow-up", payload, return_response=True)
        
        if response and response.status_code == 200:
            content = response.json().get("content")
            app.logger.info(f"Successfully generated follow-up for lead {lead_id}")
            return content
        else:
            app.logger.error(f"Failed to generate follow-up content for lead {lead_id}")
            return None
            
    except Exception as e:
        app.logger.error(f"Error generating follow-up content: {str(e)}", exc_info=True)
        return None
#-------------------------------------------------------------------------------------------------------------------------------------------------


@app.route("/process_follow_ups", methods=["GET"])
def process_follow_ups():
    # Check for secret token (similar to your /process endpoint)
    token = request.args.get("token")
    if token != os.environ.get("PROCESS_SECRET_TOKEN"):
        return jsonify({"error": "Unauthorized"}), 401
    
    try:
        # Get due follow-ups
        now = datetime.now(timezone.utc).isoformat()
        due_follow_ups = supabase.table("lead_follow_ups") \
            .select("*, leads(*)") \
            .lte("scheduled_at", now) \
            .eq("status", "pending") \
            .execute().data
        
        results = {"processed": [], "failed": []}
        
        for follow_up in due_follow_ups:
            try:
                # Generate content using AI
                if generate_follow_up_content(follow_up["lead_id"], follow_up["sequence_step"]):
                    # Update status
                    supabase.table("lead_follow_ups") \
                        .update({"status": "processed", "processed_at": now}) \
                        .eq("id", follow_up["id"]) \
                        .execute()
                    results["processed"].append(follow_up["id"])
                else:
                    supabase.table("lead_follow_ups") \
                        .update({"status": "failed", "processed_at": now}) \
                        .eq("id", follow_up["id"]) \
                        .execute()
                    results["failed"].append(follow_up["id"])
                    
            except Exception as e:
                app.logger.error(f"Error processing follow-up {follow_up['id']}: {str(e)}")
                results["failed"].append(follow_up["id"])
        
        return jsonify(results), 200
        
    except Exception as e:
        app.logger.error(f"Error in process_follow_ups: {str(e)}")
        return jsonify({"error": str(e)}), 500


#----------------------------------------------------------------------------------------------------------------------------------------
# Add to app.py

@app.route("/api/generate-complete-kit", methods=["POST"])
def generate_complete_kit():
    """Generate a complete closing kit with all document types"""
    data = request.get_json()
    ip = request.remote_addr
    
    # Check rate limits
    if (ip not in demo_rate_limits or 
        'kits' not in demo_rate_limits[ip] or 
        demo_rate_limits[ip]['kits'] <= 0):
        
        return jsonify({"error": "Closing kit limit exceeded"}), 429
    
    try:
        # Decrement the limit
        demo_rate_limits[ip]['kits'] -= 1
        
        # Generate all document types
        docs = []
        templates = [
            ("loi_template.docx", "LOI"),
            ("psa_template.docx", "PSA"),
            ("purchase_offer_template.docx", "PURCHASE_OFFER"),
            ("agency_disclosure_template.docx", "AGENCY_DISCLOSURE"),
            ("real_estate_purchase_template.docx", "REAL_ESTATE_PURCHASE"),
            ("lease_template.docx", "LEASE"),
            ("seller_disclosure_template.docx", "SELLER_DISCLOSURE"),
        ]
        
        # Create temporary directory for documents
        import tempfile
        import uuid
        tmpdir = tempfile.mkdtemp()
        
        for template_name, prefix in templates:
            try:
                tpl = DocxTemplate(f"templates/transaction_autopilot/{template_name}")
                
                # Map form data to template variables
                template_data = map_form_data_to_template(data, prefix.lower())
                tpl.render(template_data)
                
                out_name = f"{prefix}_{data.get('id', 'demo')}_{uuid.uuid4().hex[:6]}.docx"
                out_path = os.path.join(tmpdir, out_name)
                tpl.save(out_path)
                docs.append(out_path)
            except Exception as e:
                app.logger.error(f"Error generating {template_name}: {str(e)}")
                continue
        
        # Bundle into ZIP
        zip_io = BytesIO()
        with zipfile.ZipFile(zip_io, "w") as zf:
            for doc_path in docs:
                zf.write(doc_path, arcname=os.path.basename(doc_path))
        
        zip_io.seek(0)
        
        # Clean up temporary files
        for doc_path in docs:
            try:
                os.remove(doc_path)
            except:
                pass
                
        try:
            os.rmdir(tmpdir)
        except:
            pass
        
        # Return the ZIP file
        return send_file(
            zip_io,
            as_attachment=True,
            download_name=f"complete_closing_kit_{data.get('id', 'demo')}.zip",
            mimetype="application/zip"
        )
        
    except Exception as e:
        app.logger.error(f"Error generating closing kit: {str(e)}")
        return jsonify({"error": str(e)}), 500

def map_form_data_to_template(form_data, doc_type):
    """Map form data to appropriate template variables based on document type"""
    mapped_data = form_data.copy()
    
    # Add common mappings
    mapped_data['transaction_id'] = form_data.get('id', '')
    mapped_data['current_date'] = datetime.now().strftime('%B %d, %Y')
    
    # Document-specific mappings
    if doc_type == 'loi':
        mapped_data['letter_date'] = datetime.now().strftime('%B %d, %Y')
        mapped_data['buyer_signature'] = form_data.get('buyer_signature', '')
        mapped_data['seller_signature'] = form_data.get('seller_signature', '')
    
    elif doc_type == 'psa':
        mapped_data['effective_date'] = form_data.get('agreement_date', '')
        mapped_data['closing_date'] = form_data.get('closing_date', '')
        mapped_data['purchase_price'] = f"${float(form_data.get('purchase_price', 0)):,.2f}"
    
    elif doc_type == 'lease':
        mapped_data['lease_term'] = form_data.get('rent_type', '')
        mapped_data['monthly_rent'] = f"${float(form_data.get('agreed_rent', 0)):,.2f}"
        mapped_data['security_deposit'] = f"${float(form_data.get('deposit_amount', 0)):,.2f}"
    
    return mapped_data

# Add a route to check rate limit status
@app.route("/rate_limit_status")
def rate_limit_status():
    ip = request.remote_addr
    now = datetime.now()
    
    # Check and reset limits if needed (same logic as decorator)
    for resource in ['emails', 'kits', 'leads']:
        if resource == 'emails':
            if (now - demo_rate_limits[ip][resource]['last_reset']).days >= 1:
                demo_rate_limits[ip][resource]['remaining'] = 20
                demo_rate_limits[ip][resource]['last_reset'] = now
        else:
            if (now - demo_rate_limits[ip][resource]['last_reset']).days >= 30:
                if resource == 'kits':
                    demo_rate_limits[ip][resource]['remaining'] = 20
                else:
                    demo_rate_limits[ip][resource]['remaining'] = 25
                demo_rate_limits[ip][resource]['last_reset'] = now
    
    return jsonify({
        'emails_remaining': demo_rate_limits[ip]['emails']['remaining'],
        'kits_remaining': demo_rate_limits[ip]['kits']['remaining'],
        'leads_remaining': demo_rate_limits[ip]['leads']['remaining'],
        'emails_reset': (demo_rate_limits[ip]['emails']['last_reset'] + timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S'),
        'kits_reset': (demo_rate_limits[ip]['kits']['last_reset'] + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S'),
        'leads_reset': (demo_rate_limits[ip]['leads']['last_reset'] + timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
    })
#-----------------------------------------------------------------------------------------------------------------------------------------
  
# ── Final entry point ──
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 10000)))
